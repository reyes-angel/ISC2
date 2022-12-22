from config import Config, Meeting
from config import Officer
from datetime import datetime as dt
from datetime import timedelta as td
from numpy import NaN, datetime64, float64
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from os import listdir
from os.path import isfile, join
import datetime
import io
import numpy as np
import os
import pandas as pd
import re

# Zoom "meetings" do not product the same attendance report file format that is created for Zoom "webinars".  
# We should review why this is the case, maybe there is something we can do about it in the Zoom configuration.
# The primary function of this application was orignally built around the output produced for Zoom webinars.
# Later, to increse meeting enagement the chapter president made the decision to convert from webinar to meetings. 
# Rather than have to rewrite the application, I've modified it to process both types of files. 
# To process Webinar files call the "ProcessWebinarFile" function
# To process Meeting files call the "ProcessMeetingFile" function
class ReportProcessor:

    def __init__(self, configuration):
        self.config = configuration
        self.meeting = Meeting()

    def ProcessWebinarFile(self, fileName):
        in_file_name = self.config.InputPath + fileName
        io_dict = self._ReadWebinarReportFile(in_file_name)
        df, meeting_date = self._CreateWebinarDataFrame(io_dict)
        outputs = self._ProcessFrame(df, meeting_date)
        return outputs

    def ProcessMeetingFile(self, registrationFileName, attendanceFileName):
        registration_file = self.config.InputPath + registrationFileName
        attendance_file = self.config.InputPath + attendanceFileName
        
        df_registration = pd.read_csv(registration_file, header=0)
        df_registration.rename(columns={"For CPE credit provide (ISC)2 Membership ID:": "(ISC)2 Certification:"},  inplace=True)
        df_registration.drop(["Country/Region", "Unnamed: 8"], axis=1, inplace=True)
        df_registration['Email'] = df_registration['Email'].str.lower()

        df_attendance = pd.read_csv(attendance_file, header=0)
        df_attendance.rename(columns={"Name (Original Name)": "User Name (Original Name)", "Guest": "Is Guest", "User Email": "Email", "Duration (Minutes)": "Time in Session (minutes)"}, inplace=True)
        df_attendance.drop(["Recording Consent", "In Waiting Room"], axis=1, inplace=True)
        df_attendance["Attended"] = "Yes"
        df_attendance['Email'] = df_attendance['Email'].str.lower()

        df_combined = df_registration.merge(df_attendance, how="outer", on="Email", indicator=True, validate="one_to_many")
        df_combined["Attended"] = df_combined.apply(lambda x: "No" if  x["_merge"] == "left_only" else "Yes", axis=1)
        df_combined["Country/Region Name"] = df_combined.apply(lambda x: "United States" if  pd.isna(x["Country/Region Name"]) else x["Country/Region Name"], axis=1)
        
        df_combined.drop(["_merge"], axis=1, inplace=True)
        df_combined["Type"] = df_combined.apply(lambda x: "Host" if  x["Email"] == "president@isc2-siliconvalley-chapter.org" else "Attendee", axis=1)

        meeting_date = pd.to_datetime(df_attendance.iloc[0]["Join Time"])
        outputs = self._ProcessFrame(df_combined, meeting_date)
        return outputs

    def _ProcessFrame(self, processFrame, meetingDate):
        meeting_info = self.meeting.GetInformation(meetingDate)
        meeting_topic = meeting_info["Meeting_Topic"]
        cpe_group = meeting_info["CPE_Group"]
        cpe_domain = meeting_info["CPE_Domain"]

        self._ProcessDataFrame(processFrame, meetingDate.date(), meeting_topic)
        
        processFrame.drop(["Join Time", "Leave Time", "Minutes Attended", "Registration Time", "Approval Status"], axis=1, inplace=True)
        processFrame.drop_duplicates(keep="first", subset=["Email"], inplace=True)
        processFrame.sort_values(by=["Member First Name", "Member Last Name"], inplace=True, ignore_index=True)

        metrics_file_name = self.config.OutputPath + "{0}_metrics.csv".format(meetingDate.strftime("%Y%m%d"))
        attendee_file_name = self.config.OutputPath + "{0}_attendee_report.csv".format(meetingDate.strftime("%Y%m%d"))
        cpe_file_name = self.config.OutputPath  + "{0}_{1}_ISC2SiliconValleyChapter_CPE.xlsx".format(meetingDate.strftime("%b"), meetingDate.strftime("%Y"))
        
        # Export the Metrics File
        processFrame[["Attended", "Region", "Type", "CPE Qualifying Minutes", "# CPEs", "Date of Activity", "(ISC)2 Member #", "Member First Name", "Member Last Name", "Email"]].to_csv(metrics_file_name, index=False)

        # Export the Attendee File, used in creating/sending CPE certificates.  Filter to only those with an (ISC)2 Member # and who earned more than 0 CPE's
        processFrame.loc[(processFrame["# CPEs"] > 0) & (processFrame["(ISC)2 Member #"]).notna(), ["Date of Activity", "User Name", 'Email', "# CPEs", "(ISC)2 Member #"]].to_csv(attendee_file_name, index=False)


        # Export the CPE data into the template file
        # This process is a little more involved than just saving the results of a data frame to a csv because 
        # (ISC)2 wants the data to be sent as an excel file, formatted in a way that works for them to operate at scale.
        wb = load_workbook(self.config.CPETemplatePath)
        ws = wb["Chapter-CPE Submission"]

        for officer in self.config.Officers:
            if officer.Role == "Secretary":
                ws["A3"] = officer.UserName
                ws["B3"] = officer.PhoneNumber
                ws["C3"] = officer.ChapterEmail
                break

        ws["D3"] = self.config.ChapterName
        ws["E3"] = cpe_group
        ws["F3"] = cpe_domain

        #Filter to only those with an (ISC)2 Member # and who earned more than 0 CPE's
        df_cpe = processFrame.loc[(processFrame["# CPEs"] > 0) & (processFrame["(ISC)2 Member #"]).notna(), ["(ISC)2 Member #", "Member First Name", "Member Last Name", "Title of Meeting", "# CPEs", "Date of Activity"]]
        
        for r in dataframe_to_rows(df_cpe, index=False, header=False):
            ws.append(r)

        wb.save(cpe_file_name)

        outputs = {
            "attendees": attendee_file_name,
            "metrics": metrics_file_name,
            "cpe": cpe_file_name,
            "meeting_date": meetingDate
        }
        return outputs

    def _ReadWebinarReportFile(self, fileName):
        table_names = ["Host Details", "Panelist Details", "Attendee Details", "Other Attended", "Report Generated"]

        io_host = io.StringIO()
        io_panelist = io.StringIO()
        io_attendee = io.StringIO()
        io_other = io.StringIO()
        io_report = io.StringIO()
        attendee_type = "None"

        with open(fileName, "r") as f_in:
            
            for line in f_in.readlines():
                if line.startswith(table_names[0]):
                    attendee_type = "Host"
                    continue
                elif line.startswith(table_names[1]):
                    attendee_type = "Panelist"
                    continue
                elif line.startswith(table_names[2]):
                    attendee_type = "Attendee"
                    continue
                elif line.startswith(table_names[3]):
                    attendee_type = "Other"
                    continue
                elif line.startswith(table_names[4]):
                    attendee_type = "Report"
                    continue

                line = line.replace("--", "").strip().rstrip(",") + "\n"
                if attendee_type == "Host":
                    io_host.write(line)
                elif attendee_type == "Attendee":
                    io_attendee.write(line)
                elif attendee_type == "Panelist":
                    io_panelist.write(line)
                elif attendee_type == "Other":
                    io_other.write(line)
                elif attendee_type == "Report":
                    io_report.write(line)

        io_report.seek(0)
        io_host.seek(0)
        io_attendee.seek(0)
        io_panelist.seek(0)
        io_other.seek(0)

        io_dict ={
            "Report": io_report,
            "Host": io_host,
            "Attendee": io_attendee,
            "Panelist": io_panelist,
            "Other": io_other
        }

        return io_dict

    # Returns a tuple (data frame, meeting date)
    def _CreateWebinarDataFrame(self, ioDictionary):
        df_host = pd.read_csv(ioDictionary["Host"], header=0)
        df_host["Type"] = "Host"
        # We weren't able to reliably get the meeting date from the Report table because the "Actual Start Time" value is sometimes the day before the meeting!
        # We'll get it from the Host table.
        meeting_date = pd.to_datetime(df_host.iloc[0]["Join Time"])

        df_attendee = pd.read_csv(ioDictionary["Attendee"], header=0)
        df_attendee["Type"] = "Attendee"

        df_panelist = pd.read_csv(ioDictionary["Panelist"], header=0)
        df_panelist["Type"] = "Panelist"
        
        # It is common for there to not be any attendee's listed in the "Other" table
        if len(ioDictionary["Other"].getvalue()) > 0:
            df_other = pd.read_csv(ioDictionary["Other"], header=0)
            df_other.rename(columns={"User Name": "User Name (Original Name)"},  inplace=True)
            df_other["Type"] = "Other"
            df_other["Email"] = df_other["User Name (Original Name)"].map(str)
            df_other["First Name"] = "Unknown"
            df_other["Last Name"] = "Unknown"
            df_other["Attended"] = "Yes"
        else:
            df_other = pd.DataFrame()

        df = pd.concat([df_host, df_panelist, df_attendee, df_other], ignore_index=True)
        return df, meeting_date

    def _ProcessDataFrame(self, df, meeting_date, meeting_topic):
        meeting_start = np.datetime64(dt.combine(meeting_date, datetime.time(self.config.MeetingStartHour, 00)))
        meeting_end = np.datetime64(dt.combine(meeting_date, datetime.time(self.config.MeetingEndHour, 00)))

        df.rename(columns={"First Name": "Member First Name", "Last Name": "Member Last Name", "User Name (Original Name)": "User Name", "Time in Session (minutes)": "Minutes Attended", "Country/Region Name": "Region", "(ISC)2 Certification:": "(ISC)2 Member #"},  inplace=True)

        # Chapter officers will likely register to attend the meeting under their personal email address and provide thier Member #.  However, when they join the meeting
        # the panelist/host email invite may have them joining under thier chapter email addres and not associate a Member #.  In this case, they would not get CPE credit for having
        # joined the call.  To avoid this, we update the chapter officers record with thier first, last, and member #.
        for officer in self.config.Officers:
            df.loc[df["Email"]== officer.ChapterEmail, ["User Name"]] = officer.UserName
            df.loc[df["Email"]== officer.ChapterEmail, ["Member First Name"]] = officer.FirstName
            df.loc[df["Email"]== officer.ChapterEmail, ["Member Last Name"]] = officer.LastName
            df.loc[df["Email"]== officer.ChapterEmail, ["(ISC)2 Member #"]] = officer.CertificationID
            df.loc[df["Email"]== officer.PersonalEmail,["Email"]] = officer.ChapterEmail

        # Some attendees fail to add only their certification ID when registering for the meeting.
        # Here we extract the numeric digits from the "(ISC)2 Member #" filed and replace the existing field value.
        # Most certification nunbers are 6 digit but this is looking for a number between 4 and 9 digits long.
        df["(ISC)2 Member #"] = df["(ISC)2 Member #"].astype(str).str.extract(r"([0-9]{4,9})")
        df["Join Time"] = pd.to_datetime(df["Join Time"])
        df["Leave Time"] = pd.to_datetime(df["Leave Time"])

        # Add new fields to the data frame
        df["# CPEs"] = 0.0
        df["CPE Qualifying Minutes"] = 0.0
        df["Date of Activity"] = dt.strftime(meeting_date, "%m/%d/%Y")
        df["Title of Meeting"] = meeting_topic
        
        # When someone has connected to the meeting more than once there are multiple entries in the report.  
        # Only the first will list an "(ISC)2 Member #" number, assuming the attendee provided one when they registered.
        # Here we backfill all of those empty values.  
        df["(ISC)2 Member #"] = df.groupby(["Email"])["(ISC)2 Member #"].transform("first")

        # Covers Attendee Use Case 1 and 2
        # Discard any record where the attendee either left before the meeting started or joined after it ended
        df.drop(df.loc[(df["Join Time"] >= meeting_end) | (df["Leave Time"] <= meeting_start)].index, axis=0, inplace=True)

        # Covers Attendee Use Case 3A
        self._DiscardInvtervals(df)

        # Covers Attendee Use Case 3B
        self._BridgeIntevals(df)

        # Covers Attendee Use Case 4
        # If the attendee joined at any point before the meeting we will consider the join time to be at the start of the meeting (typically 6:00 PM)
        df["Join Time"] = df.apply(lambda x: meeting_start if  x["Join Time"] < meeting_start else x["Join Time"], axis=1)

        # Covers Attendee Use Case 5
        # If the attendee left the meeting after the meeting ended we will consider the leave time to the at the end of the meeting (typically 8:00 PM)
        df["Leave Time"] = df.apply(lambda x: meeting_end if  x["Leave Time"] >= meeting_end else x["Leave Time"], axis=1)

        # Calculate how many minutes of the meeting were attended, per record (not yet aggregated)
        df["Minutes Attended"] = (df["Leave Time"] - df["Join Time"]).astype("timedelta64[m]")

        #Calculate CPE Qualifying Minutes and # CPEs
        df["CPE Qualifying Minutes"] = df.groupby(["Email"])["Minutes Attended"].transform("sum")

        # If the attendee attended at least 110 minutes, of the 120 minute meeting, give full credit for the entire meeting.
        df["CPE Qualifying Minutes"] = df.apply(lambda x: 120 if  x["CPE Qualifying Minutes"] >= 110 else x["CPE Qualifying Minutes"], axis=1)

        # Round CPE Qualifying Minutes to the nearest 15 minutes and then divide by 60 to find the number of CPE's to assign for attendance
        # Each 15 minutes of attendance gives the attendee .25 CPE's.  An attendee that accumulated less than 15 minutes of attendance will receive 0 CPE's 
        df["# CPEs"] = ((df["CPE Qualifying Minutes"]/15).round().astype(int) * 15)/60


    def _DiscardInvtervals(self, df):
        df.sort_values(by=["Email", "Join Time", "Leave Time"], ascending=[True, True, False], inplace=True, ignore_index=True)
        index_list_to_delete = []

        for email_address in df[df.duplicated(subset=["Email"], keep="first")]["Email"].unique():
            filter_by_email = df["Email"].eq(email_address)
            attendee_intervals = df[filter_by_email].filter(items=["Join Time", "Leave Time"], axis=1).to_dict()
            
            max_index = max(attendee_intervals["Join Time"].keys())
            min_index = min(attendee_intervals["Join Time"].keys())

            index = max_index

            while index > min_index:
                current_start = attendee_intervals["Join Time"][index]
                current_end = attendee_intervals["Leave Time"][index]
                
                compare_index = max_index if index == min_index else index - 1
                previous_start = attendee_intervals["Join Time"][compare_index]
                previous_end = attendee_intervals["Leave Time"][compare_index]

                if(current_start >= previous_start and current_end <= previous_end):
                    index_list_to_delete.append(index)
                
                index -= 1

        df.drop(index_list_to_delete, axis=0, inplace=True)


    def _BridgeIntevals(self, df):
        df.sort_values(by=["Email", "Join Time", "Leave Time"], ascending=[True, True, False], inplace=True, ignore_index=True)
        index_list_to_delete = []

        for email_address in df[df.duplicated(subset=["Email"], keep="first")]["Email"].unique():
            filter_by_email = df["Email"].eq(email_address)
            attendee_intervals = df[filter_by_email].filter(items=["Join Time", "Leave Time"], axis=1).to_dict()
            
            max_index = max(attendee_intervals["Join Time"].keys())
            min_index = min(attendee_intervals["Join Time"].keys())

            index = max_index

            while index > min_index:
                current_start = attendee_intervals["Join Time"][index]
                current_end = attendee_intervals["Leave Time"][index]
                
                compare_index = max_index if index == min_index else index - 1
                previous_end = attendee_intervals["Leave Time"][compare_index]
                
                if((current_start - td(minutes=1)) <= previous_end):
                    # Here we're updating both the dataframe and the attendee intervals dictionary
                    # The source of our comparisons is the dictionary, if we only update the dataframe 
                    # we won't have the most up to date information
                    df.loc[compare_index, ["Leave Time"]] = current_end
                    attendee_intervals["Leave Time"][compare_index] = current_end
                    index_list_to_delete.append(index)
                
                index -= 1

        df.drop(index_list_to_delete, axis=0, inplace=True)
